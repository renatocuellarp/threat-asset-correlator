# src/exportacion/exportador.py
# Módulo de exportación Excel y PDF
# Threat Asset Correlator — Renato Cuellar

import io
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, HRFlowable
)

from src.assets.modelo import Activo
from src.risk.correlador import ResultadoCorrelacion, TipoRiesgo


# ── Colores corporativos ──
COLOR_EXTREMO = "7F0000"
COLOR_ALTO = "C0392B"
COLOR_MODERADO = "E67E22"
COLOR_BAJO = "27AE60"
COLOR_HEADER = "1A1D24"


def _color_por_nivel(nivel: str) -> str:
    return {
        "EXTREMO": COLOR_EXTREMO,
        "ALTO": COLOR_ALTO,
        "MODERADO": COLOR_MODERADO,
        "BAJO": COLOR_BAJO,
    }.get(nivel, "888888")


def exportar_excel(
    activo: Activo,
    resultados: List[ResultadoCorrelacion],
    tipo_riesgo: TipoRiesgo
) -> bytes:
    wb = Workbook()

    # ── Hoja 1: Resumen ──
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color=COLOR_HEADER)
    center = Alignment(horizontal="center", vertical="center")

    # Título
    ws_resumen["A1"] = "REPORTE DE CORRELACIÓN DE AMENAZAS"
    ws_resumen["A1"].font = Font(name="Arial", bold=True, size=14)
    ws_resumen.merge_cells("A1:G1")
    ws_resumen["A1"].alignment = center

    # Metadatos
    ws_resumen["A3"] = "Activo"
    ws_resumen["B3"] = activo.nombre
    ws_resumen["A4"] = "Criticidad"
    ws_resumen["B4"] = activo.criticidad.name
    ws_resumen["A5"] = "Tipo de riesgo"
    ws_resumen["B5"] = tipo_riesgo.value
    ws_resumen["A6"] = "Fecha de análisis"
    ws_resumen["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_resumen["A7"] = "Fuente de inteligencia"
    ws_resumen["B7"] = "NVD/CVE (NIST)"
    ws_resumen["A8"] = "Total correlaciones"
    ws_resumen["B8"] = len(resultados)

    for row in range(3, 9):
        ws_resumen[f"A{row}"].font = Font(name="Arial", bold=True)

    # Resumen por nivel
    ws_resumen["A10"] = "RESUMEN POR NIVEL DE RIESGO"
    ws_resumen["A10"].font = Font(name="Arial", bold=True, size=12)

    niveles = ["EXTREMO", "ALTO", "MODERADO", "BAJO"]
    for col, nivel in enumerate(niveles, start=1):
        cell = ws_resumen.cell(row=11, column=col)
        cell.value = nivel
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=_color_por_nivel(nivel))
        cell.alignment = center

        count_cell = ws_resumen.cell(row=12, column=col)
        count_cell.value = sum(1 for r in resultados if r.nivel_riesgo() == nivel)
        count_cell.font = Font(name="Arial", size=14, bold=True)
        count_cell.alignment = center

    for col in range(1, 5):
        ws_resumen.column_dimensions[get_column_letter(col)].width = 18

    # ── Hoja 2: Correlaciones ──
    ws_corr = wb.create_sheet("Correlaciones")

    headers = [
        "CVE ID", "Severidad", "Score CVSS",
        "Probabilidad", "Impacto", "Score Total", "Nivel de Riesgo",
        "Fecha Publicación", "Descripción"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws_corr.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row, r in enumerate(resultados, start=2):
        ws_corr.cell(row=row, column=1).value = r.cve.id
        ws_corr.cell(row=row, column=2).value = r.cve.severidad
        ws_corr.cell(row=row, column=3).value = round(r.cve.score_cvss, 1)
        ws_corr.cell(row=row, column=4).value = r.probabilidad.name
        ws_corr.cell(row=row, column=5).value = r.impacto.name
        ws_corr.cell(row=row, column=6).value = r.score_total
        ws_corr.cell(row=row, column=7).value = r.nivel_riesgo()
        ws_corr.cell(row=row, column=8).value = r.cve.fecha_publicacion
        ws_corr.cell(row=row, column=9).value = r.cve.descripcion

        # Color por nivel de riesgo
        nivel_cell = ws_corr.cell(row=row, column=7)
        nivel_cell.fill = PatternFill(
            "solid",
            start_color=_color_por_nivel(r.nivel_riesgo())
        )
        nivel_cell.font = Font(name="Arial", color="FFFFFF", bold=True)

        for col in range(1, 10):
            ws_corr.cell(row=row, column=col).font = Font(name="Arial")

    # Anchos de columna
    anchos = [15, 12, 12, 14, 14, 12, 16, 18, 60]
    for col, ancho in enumerate(anchos, start=1):
        ws_corr.column_dimensions[get_column_letter(col)].width = ancho

    # ── Hoja 3: Justificación técnica ──
    ws_just = wb.create_sheet("Justificación técnica")

    just_headers = ["CVE ID", "Nivel de Riesgo", "Justificación"]
    for col, header in enumerate(just_headers, start=1):
        cell = ws_just.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row, r in enumerate(resultados, start=2):
        ws_just.cell(row=row, column=1).value = r.cve.id
        ws_just.cell(row=row, column=2).value = r.nivel_riesgo()
        ws_just.cell(row=row, column=3).value = r.justificacion

        nivel_cell = ws_just.cell(row=row, column=2)
        nivel_cell.fill = PatternFill(
            "solid",
            start_color=_color_por_nivel(r.nivel_riesgo())
        )
        nivel_cell.font = Font(name="Arial", color="FFFFFF", bold=True)

        for col in range(1, 4):
            cell = ws_just.cell(row=row, column=col)
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_just.column_dimensions["A"].width = 18
    ws_just.column_dimensions["B"].width = 16
    ws_just.column_dimensions["C"].width = 80

    # Exportar a bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def exportar_pdf(
    activo: Activo,
    resultados: List[ResultadoCorrelacion],
    tipo_riesgo: TipoRiesgo
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch
    )

    styles = getSampleStyleSheet()
    story = []

    # ── Título ──
    titulo_style = ParagraphStyle(
        "Titulo",
        parent=styles["Title"],
        fontSize=18,
        spaceAfter=6,
        textColor=colors.HexColor("#e0e0e0"),
        backColor=colors.HexColor("#1a1d24"),
        borderPadding=10,
    )
    story.append(Paragraph("Reporte de Correlación de Amenazas", titulo_style))
    story.append(Spacer(1, 12))
    story.append(HRFlowable(width="100%", color=colors.HexColor("#c0392b")))
    story.append(Spacer(1, 12))

    # ── Metadatos ──
    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["Normal"],
        fontSize=10,
        spaceAfter=4,
    )

    metadatos = [
        ["Activo", activo.nombre],
        ["Criticidad", activo.criticidad.name],
        ["Tipo de riesgo", tipo_riesgo.value],
        ["Fecha de análisis", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Fuente", "NVD/CVE (NIST)"],
        ["Total correlaciones", str(len(resultados))],
    ]

    meta_table = Table(metadatos, colWidths=[2 * inch, 4 * inch])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f0f0")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f9f9f9")]),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 20))

    # ── Resumen por nivel ──
    story.append(Paragraph("Resumen por nivel de riesgo", styles["Heading2"]))
    story.append(Spacer(1, 8))

    niveles = ["EXTREMO", "ALTO", "MODERADO", "BAJO"]
    resumen_data = [niveles, [
        str(sum(1 for r in resultados if r.nivel_riesgo() == n))
        for n in niveles
    ]]

    resumen_table = Table(resumen_data, colWidths=[1.5 * inch] * 4)
    resumen_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, 1), 16),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#7f0000")),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#c0392b")),
        ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#e67e22")),
        ("BACKGROUND", (3, 0), (3, 0), colors.HexColor("#27ae60")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHTS", (0, 0), (-1, 0), 25),
        ("ROWHEIGHTS", (0, 1), (-1, 1), 35),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.white),
    ]))
    story.append(resumen_table)
    story.append(Spacer(1, 20))

    # ── Tabla de correlaciones ──
    story.append(Paragraph("Detalle de correlaciones", styles["Heading2"]))
    story.append(Spacer(1, 8))

    corr_headers = ["CVE ID", "Severidad", "CVSS", "Prob.", "Impacto", "Score", "Nivel"]
    corr_data = [corr_headers] + [
        [
            r.cve.id,
            r.cve.severidad,
            str(round(r.cve.score_cvss, 1)),
            r.probabilidad.name,
            r.impacto.name,
            str(r.score_total),
            r.nivel_riesgo(),
        ]
        for r in resultados
    ]

    corr_table = Table(
        corr_data,
        colWidths=[1.2*inch, 0.9*inch, 0.6*inch, 0.8*inch, 0.9*inch, 0.6*inch, 0.9*inch]
    )

    table_style = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1d24")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
    ]

    nivel_colores = {
        "EXTREMO": "#7f0000",
        "ALTO": "#c0392b",
        "MODERADO": "#e67e22",
        "BAJO": "#27ae60",
    }

    for row_idx, r in enumerate(resultados, start=1):
        color = nivel_colores.get(r.nivel_riesgo(), "#888888")
        table_style.append(
            ("BACKGROUND", (6, row_idx), (6, row_idx), colors.HexColor(color))
        )
        table_style.append(
            ("TEXTCOLOR", (6, row_idx), (6, row_idx), colors.white)
        )
        table_style.append(
            ("FONTNAME", (6, row_idx), (6, row_idx), "Helvetica-Bold")
        )

    corr_table.setStyle(TableStyle(table_style))
    story.append(corr_table)
    story.append(Spacer(1, 20))

    # ── Justificación técnica ──
    story.append(Paragraph("Justificación técnica", styles["Heading2"]))
    story.append(Spacer(1, 8))

    just_style = ParagraphStyle(
        "Just",
        parent=styles["Normal"],
        fontSize=8,
        spaceAfter=6,
        leading=12,
    )

    for r in resultados:
        nivel = r.nivel_riesgo()
        color = colors.HexColor(nivel_colores.get(nivel, "#888888"))
        story.append(Paragraph(
            f"<b>{r.cve.id}</b> — <font color='#{nivel_colores.get(nivel,'888888')[1:]}'>{nivel}</font>",
            styles["Heading3"]
        ))
        story.append(Paragraph(r.justificacion, just_style))
        story.append(Paragraph(f"<i>Descripción: {r.cve.descripcion}</i>", just_style))
        story.append(HRFlowable(width="100%", color=colors.HexColor("#dddddd")))
        story.append(Spacer(1, 6))

    doc.build(story)
    return buffer.getvalue()